from __future__ import annotations

import argparse
import csv
import os
import platform
import re
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable, Optional

import psutil
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


RESULT_HEADERS = [
    "test_software",
    "test_case",
    "start_time",
    "end_time",
    "duration_seconds",
]
PROGRESS_PATTERN = (
    r"^\s*(?P<percent>\d+)% done\s+frames:\s+(?P<frames>\d+)\s+elapsed:\s+"
    r"(?P<elapsed>\d{2}:\d{2}:\d{2},\d{3})\s*$"
)
TIMESTAMP_PATTERN = r"^\[[^\]]+\]\s(?P<clock>\d{2}:\d{2}:\d{2})-(?P<millis>\d{3})"
COUNTED_OUTPUT_PATTERN = re.compile(r"^(?P<base>.+?)_run(?P<index>\d+)$", re.IGNORECASE)
AVIDEMUX_SAVE_LOOP_MARKERS = ("[FF] Saving",)
AVIDEMUX_END_MARKERS = ("End of flush",)
GENERIC_DEVICE_VALUES = {
    "",
    "default string",
    "system product name",
    "system version",
    "to be filled by o.e.m.",
    "to be filled by oem",
    "not available",
    "unknown",
}


def expand_path(raw_path: str) -> Path:
    normalized = re.sub(
        r"\$env:(?P<name>[A-Za-z_][A-Za-z0-9_]*)",
        lambda match: os.environ.get(match.group("name"), match.group(0)),
        raw_path,
        flags=re.IGNORECASE,
    )
    expanded = os.path.expandvars(os.path.expanduser(normalized))
    return Path(expanded).resolve(strict=False)


def format_bytes(value: int) -> str:
    units = ["B", "KB", "MB", "GB", "TB"]
    size = float(value)
    for unit in units:
        if size < 1024 or unit == units[-1]:
            return f"{size:.1f} {unit}"
        size /= 1024
    return f"{value} B"


def parse_iso_datetime(raw_value: str) -> Optional[datetime]:
    if not raw_value:
        return None
    try:
        return datetime.fromisoformat(raw_value)
    except ValueError:
        return None


def parse_elapsed_seconds(raw_value: str) -> Optional[float]:
    try:
        hour_value = int(raw_value[0:2])
        minute_value = int(raw_value[3:5])
        second_value = int(raw_value[6:8])
        millisecond_value = int(raw_value[9:12])
    except (ValueError, IndexError):
        return None
    return hour_value * 3600 + minute_value * 60 + second_value + millisecond_value / 1000.0


def format_datetime_or_empty(value: Optional[datetime]) -> str:
    if value is None:
        return ""
    return value.isoformat(timespec="milliseconds")


def normalize_log_line(line: str) -> str:
    return line.lstrip("\ufeff").rstrip("\r\n")


def read_csv_rows(csv_path: Path) -> list[dict[str, str]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def extract_run_label_from_csv_path(csv_path: Path) -> str:
    match = COUNTED_OUTPUT_PATTERN.fullmatch(csv_path.stem)
    if not match:
        return ""
    return f"run{int(match.group('index')):03d}"


def extract_named_case_from_csv_path(csv_path: Path, software: str) -> str:
    if not software:
        return ""

    match = COUNTED_OUTPUT_PATTERN.fullmatch(csv_path.stem)
    if match:
        base_stem = match.group("base")
    else:
        base_stem = csv_path.stem

    software_prefix = f"{software}_"
    if not base_stem.lower().startswith(software_prefix.lower()):
        return ""

    named_case = base_stem[len(software_prefix) :].strip("_")
    return named_case


def iter_input_csv_paths(inputs: Iterable[str]) -> list[Path]:
    csv_paths: list[Path] = []
    for raw_input in inputs:
        candidate = expand_path(raw_input)
        if candidate.is_dir():
            csv_paths.extend(sorted(path for path in candidate.rglob("*.csv") if path.is_file()))
        elif candidate.is_file():
            csv_paths.append(candidate)
    unique_paths: list[Path] = []
    seen = set()
    for path in csv_paths:
        if path not in seen:
            seen.add(path)
            unique_paths.append(path)
    return unique_paths


def resolve_output_path(raw_output: str | Path) -> Path:
    if isinstance(raw_output, Path):
        return raw_output.resolve(strict=False)
    return expand_path(raw_output)


def default_xlsx_output_path(csv_path: str | Path) -> Path:
    resolved_csv_path = resolve_output_path(csv_path)
    return resolved_csv_path.with_suffix(".xlsx")


def run_powershell_lines(script: str, *, timeout: int = 10) -> list[str]:
    command = ["powershell", "-NoProfile", "-Command", script]
    try:
        result = subprocess.run(command, capture_output=True, text=True, timeout=timeout, check=False)
    except (OSError, subprocess.SubprocessError):
        return []
    if result.returncode != 0:
        return []
    return [line.strip() for line in result.stdout.splitlines() if line.strip()]


def get_registry_value(path: str, property_name: str) -> str:
    script = f"(Get-ItemProperty '{path}').{property_name}"
    values = run_powershell_lines(script)
    return values[0] if values else ""


def is_meaningful_device_value(value: str) -> bool:
    return value.strip().lower() not in GENERIC_DEVICE_VALUES


def get_device_model() -> str:
    bios_path = r"HKLM:\HARDWARE\DESCRIPTION\System\BIOS"
    candidates = [
        get_registry_value(bios_path, "SystemFamily"),
        get_registry_value(bios_path, "SystemSKU"),
        get_registry_value(bios_path, "SystemProductName"),
        get_registry_value(bios_path, "BaseBoardProduct"),
    ]
    for candidate in candidates:
        if is_meaningful_device_value(candidate):
            return candidate
    return platform.node().strip()


def get_cpu_name() -> str:
    cpu_name = get_registry_value(
        r"HKLM:\HARDWARE\DESCRIPTION\System\CentralProcessor\0",
        "ProcessorNameString",
    )
    if cpu_name:
        return cpu_name
    processor = platform.processor().strip()
    return processor


def get_gpu_names() -> str:
    names = run_powershell_lines(
        "(Get-ItemProperty 'HKLM:\\SYSTEM\\CurrentControlSet\\Control\\Video\\*\\0000').DriverDesc",
    )
    unique_names = list(dict.fromkeys(names))
    return "; ".join(unique_names)


def get_disk_summary() -> str:
    disk_lines: list[str] = []
    total_capacity = 0
    seen_mountpoints = set()
    for partition in psutil.disk_partitions(all=False):
        mountpoint = partition.mountpoint
        if mountpoint in seen_mountpoints:
            continue
        seen_mountpoints.add(mountpoint)
        try:
            usage = psutil.disk_usage(mountpoint)
        except OSError:
            continue
        total_capacity += usage.total
        drive_name = mountpoint.rstrip("\\/")
        disk_lines.append(f"{drive_name} total={format_bytes(usage.total)}")

    if not disk_lines:
        return ""
    return f"total={format_bytes(total_capacity)}; " + "; ".join(disk_lines)


def get_ram_total() -> str:
    return format_bytes(psutil.virtual_memory().total)


def get_windows_version() -> str:
    windows_name = platform.system().strip() or "Windows"
    release = platform.release().strip()
    version = platform.version().strip()
    return f"{windows_name} {release} ({version})"


def collect_device_info() -> list[tuple[str, str]]:
    gpu_value = get_gpu_names()
    return [
        ("device_name", get_device_model()),
        ("windows_system_version", get_windows_version()),
        ("cpu", get_cpu_name()),
        ("gpu", gpu_value),
        ("mem_ram", get_ram_total()),
        ("disk", get_disk_summary()),
    ]


def infer_avidemux_timestamps(source_path: Path, duration_seconds: float) -> tuple[Optional[datetime], Optional[datetime]]:
    if not source_path.exists():
        return None, None

    lines = source_path.read_text(encoding="utf-8", errors="ignore").splitlines()
    target_index: Optional[int] = None
    file_date = datetime.fromtimestamp(source_path.stat().st_mtime).astimezone().date()

    for index, line in enumerate(lines):
        match = re.search(PROGRESS_PATTERN, normalize_log_line(line))
        if not match:
            continue
        if int(match.group("percent")) != 100:
            continue
        elapsed_value = parse_elapsed_seconds(match.group("elapsed"))
        if elapsed_value is None:
            continue
        if abs(elapsed_value - duration_seconds) <= 0.01:
            target_index = index

    if target_index is None:
        return None, None

    def parse_timestamped_line(raw_line: str) -> Optional[datetime]:
        timestamp_match = re.search(TIMESTAMP_PATTERN, normalize_log_line(raw_line))
        if not timestamp_match:
            return None
        clock_value = timestamp_match.group("clock")
        millis_value = int(timestamp_match.group("millis"))
        hour_value, minute_value, second_value = (int(value) for value in clock_value.split(":"))
        return datetime(
            year=file_date.year,
            month=file_date.month,
            day=file_date.day,
            hour=hour_value,
            minute=minute_value,
            second=second_value,
            microsecond=millis_value * 1000,
            tzinfo=datetime.now().astimezone().tzinfo,
        )

    def find_latest_timestamp(start_index: int, end_index: int) -> Optional[datetime]:
        for index in range(end_index - 1, start_index - 1, -1):
            timestamp = parse_timestamped_line(lines[index])
            if timestamp is not None:
                return timestamp
        return None

    save_loop_index: Optional[int] = None
    for index in range(0, target_index + 1):
        if any(marker in normalize_log_line(lines[index]) for marker in AVIDEMUX_SAVE_LOOP_MARKERS):
            save_loop_index = index

    explicit_end_index: Optional[int] = None
    for index in range(target_index, -1, -1):
        if any(marker in normalize_log_line(lines[index]) for marker in AVIDEMUX_END_MARKERS):
            explicit_end_index = index
            break

    end_time = (
        find_latest_timestamp(0, explicit_end_index + 1)
        if explicit_end_index is not None
        else find_latest_timestamp(0, target_index + 1)
    )
    if end_time is None:
        end_time = find_latest_timestamp(target_index + 1, len(lines))
    if end_time is None:
        return None, None

    start_time = (
        find_latest_timestamp(0, save_loop_index + 1)
        if save_loop_index is not None
        else None
    )
    if start_time is None:
        start_time = end_time - timedelta(seconds=duration_seconds)
    return start_time, end_time


def build_result_row(
    csv_path: Path,
    row: dict[str, str],
    *,
    default_test_case: str,
) -> dict[str, object]:
    software = row.get("software", "").strip()
    duration_seconds = row.get("duration_seconds", "").strip() or row.get("computed_duration_seconds", "").strip()
    numeric_duration = float(duration_seconds) if duration_seconds else None
    start_time = parse_iso_datetime(row.get("started_at", "").strip())
    end_time = parse_iso_datetime(row.get("ended_at", "").strip())

    if (start_time is None or end_time is None) and software == "avidemux" and numeric_duration is not None:
        source_path = Path(row.get("source_path", ""))
        inferred_start, inferred_end = infer_avidemux_timestamps(source_path, numeric_duration)
        start_time = start_time or inferred_start
        end_time = end_time or inferred_end

    run_label = extract_run_label_from_csv_path(csv_path)
    explicit_test_case = row.get("test_case", "").strip()
    explicit_test_name = row.get("test_name", "").strip()
    inferred_test_name = extract_named_case_from_csv_path(csv_path, software)
    if explicit_test_case:
        test_case = explicit_test_case
    elif explicit_test_name and run_label:
        test_case = f"{explicit_test_name}_{run_label}"
    elif explicit_test_name:
        test_case = explicit_test_name
    elif inferred_test_name and run_label:
        test_case = f"{inferred_test_name}_{run_label}"
    elif inferred_test_name:
        test_case = inferred_test_name
    elif default_test_case and run_label:
        test_case = f"{default_test_case}_{run_label}"
    elif default_test_case:
        test_case = default_test_case
    elif run_label:
        test_case = run_label
    else:
        test_case = row.get("session_id", "").strip()

    return {
        "test_software": software,
        "test_case": test_case,
        "start_time": format_datetime_or_empty(start_time),
        "end_time": format_datetime_or_empty(end_time),
        "duration_seconds": duration_seconds,
    }


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 60)


def build_workbook(device_info: list[tuple[str, str]], result_rows: list[dict[str, object]]) -> Workbook:
    workbook = Workbook()
    device_sheet = workbook.active
    device_sheet.title = "DeviceInfo"
    result_sheet = workbook.create_sheet("TestResults")

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    device_sheet.append(["field", "value"])
    for cell in device_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    for field_name, field_value in device_info:
        device_sheet.append([field_name, field_value])
    for row in device_sheet.iter_rows():
        for cell in row:
            cell.alignment = wrap_alignment
    autosize_columns(device_sheet)

    result_sheet.append(RESULT_HEADERS)
    for cell in result_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    for result_row in result_rows:
        result_sheet.append([result_row[header] for header in RESULT_HEADERS])
    for row in result_sheet.iter_rows():
        for cell in row:
            cell.alignment = wrap_alignment
    result_sheet.freeze_panes = "A2"
    autosize_columns(result_sheet)
    return workbook


def collect_result_rows_from_csv_paths(
    csv_paths: Iterable[Path],
    *,
    default_test_case: str = "",
) -> list[dict[str, object]]:
    all_result_rows: list[dict[str, object]] = []
    for csv_path in csv_paths:
        for row in read_csv_rows(csv_path):
            all_result_rows.append(
                build_result_row(
                    csv_path,
                    row,
                    default_test_case=default_test_case,
                )
            )
    return all_result_rows


def generate_xlsx_report(
    inputs: Iterable[str | Path],
    output: str | Path,
    *,
    default_test_case: str = "",
) -> tuple[int, int, Path]:
    normalized_inputs = [str(item) if isinstance(item, Path) else item for item in inputs]
    csv_paths = iter_input_csv_paths(normalized_inputs)
    if not csv_paths:
        raise ValueError("No csv files were found from the provided inputs.")

    all_result_rows = collect_result_rows_from_csv_paths(
        csv_paths,
        default_test_case=default_test_case,
    )
    output_path = resolve_output_path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = build_workbook(collect_device_info(), all_result_rows)
    workbook.save(output_path)
    return len(csv_paths), len(all_result_rows), output_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate an xlsx report from monitoring csv files.")
    parser.add_argument(
        "--input",
        nargs="+",
        required=True,
        help="One or more csv files or directories containing csv files.",
    )
    parser.add_argument(
        "--output",
        default="results/test_report.xlsx",
        help="Target xlsx report path. Default: results/test_report.xlsx",
    )
    parser.add_argument(
        "--default-test-case",
        default="",
        help="Fallback test case name when the source row does not provide one.",
    )
    return parser


def main(argv: Optional[list[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    try:
        csv_count, row_count, output_path = generate_xlsx_report(
            args.input,
            args.output,
            default_test_case=args.default_test_case,
        )
    except ValueError as exc:
        parser.error(str(exc))

    print(f"Wrote xlsx report: {output_path}")
    print(f"Included csv files: {csv_count}")
    print(f"Included result rows: {row_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
