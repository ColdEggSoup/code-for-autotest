from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from automation_components import SOFTWARE_SPECS, resolve_input_video
from full_test_pipeline import run_pipeline


pytestmark = [pytest.mark.workflow, pytest.mark.windows, pytest.mark.manual, pytest.mark.e2e]


def _effective_input_video(args) -> Path:
    if args.input_video:
        return Path(args.input_video).resolve(strict=False)
    return resolve_input_video(args.workload_name)


def _effective_blend_file(args) -> Path:
    if args.blend_file:
        return Path(args.blend_file).resolve(strict=False)
    default_blend = SOFTWARE_SPECS["blender"].blend_file
    assert default_blend is not None
    return default_blend.resolve(strict=False)


def _skip_if_missing_runtime_prerequisites(args) -> None:
    missing: list[str] = []

    try:
        input_video_path = _effective_input_video(args)
        if not input_video_path.exists():
            missing.append(f"input video: {input_video_path}")
    except AssertionError as exc:
        missing.append(str(exc))

    for software in args.software:
        if software == "blender":
            blend_file = _effective_blend_file(args)
            if not blend_file.exists():
                missing.append(f"blender blend file: {blend_file}")
            continue
        launch_path = SOFTWARE_SPECS[software].launch_path
        if launch_path is None or not launch_path.exists():
            missing.append(f"{software} launcher: {launch_path}")

    if missing:
        pytest.skip("Missing local runtime prerequisites for the e2e workflow: " + "; ".join(missing))


def test_baseline_pipeline_generates_xlsx_report(build_pipeline_namespace, record_property) -> None:
    args = build_pipeline_namespace(skip_turbo=True)
    _skip_if_missing_runtime_prerequisites(args)

    report_path = run_pipeline(args)
    record_property("pipeline_report", str(report_path))
    record_property("pipeline_run_root", str(report_path.parent.parent))
    workbook = load_workbook(report_path, data_only=False)
    try:
        assert "TestResults" in workbook.sheetnames
        assert "Comparison" in workbook.sheetnames
        result_rows = [row for row in workbook["TestResults"].iter_rows(min_row=2, values_only=True) if row[0]]
        assert len(result_rows) >= len(args.software)
        assert {row[0] for row in result_rows} == set(args.software)
        assert workbook["Comparison"].max_row == 1
    finally:
        workbook.close()


@pytest.mark.ai_turbo
def test_baseline_vs_ai_turbo_pipeline_generates_comparison_sheet(build_pipeline_namespace, record_property) -> None:
    args = build_pipeline_namespace()
    _skip_if_missing_runtime_prerequisites(args)

    report_path = run_pipeline(args)
    record_property("pipeline_report", str(report_path))
    record_property("pipeline_run_root", str(report_path.parent.parent))
    workbook = load_workbook(report_path, data_only=False)
    try:
        comparison_sheet = workbook["Comparison"]
        comparison_rows = [
            index
            for index in range(2, comparison_sheet.max_row + 1)
            if comparison_sheet[f"A{index}"].value
        ]
        assert len(comparison_rows) == len(args.software)
        assert {comparison_sheet[f"A{index}"].value for index in comparison_rows} == set(args.software)
        for index in comparison_rows:
            assert comparison_sheet[f"G{index}"].value == (
                f'=IF(OR(E{index}="",F{index}="",E{index}<=0),"",(E{index}-F{index})/E{index})'
            )
    finally:
        workbook.close()
