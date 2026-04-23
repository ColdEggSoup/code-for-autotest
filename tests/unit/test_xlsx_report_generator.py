from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from xlsx_report_generator import generate_xlsx_report


def _write_monitor_csv(csv_path: Path, *, software: str, test_name: str, duration_seconds: float) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=("software", "duration_seconds", "started_at", "ended_at", "test_name"),
        )
        writer.writeheader()
        writer.writerow(
            {
                "software": software,
                "duration_seconds": str(duration_seconds),
                "started_at": "2026-04-21T10:00:00.000+08:00",
                "ended_at": "2026-04-21T10:00:10.000+08:00",
                "test_name": test_name,
            }
        )


def test_generate_xlsx_report_builds_comparison_sheet_with_improvement_formula(tmp_path: Path) -> None:
    baseline_csv = tmp_path / "csv" / "shotcut_demo_baseline.csv"
    turbo_csv = tmp_path / "csv" / "shotcut_demo_turbo.csv"
    _write_monitor_csv(baseline_csv, software="shotcut", test_name="demo_baseline", duration_seconds=10.0)
    _write_monitor_csv(turbo_csv, software="shotcut", test_name="demo_turbo", duration_seconds=5.0)

    output_path = tmp_path / "report.xlsx"
    csv_count, row_count, generated_output_path = generate_xlsx_report([baseline_csv, turbo_csv], output_path)

    assert csv_count == 2
    assert row_count == 2
    assert generated_output_path == output_path

    workbook = load_workbook(generated_output_path, data_only=False)
    try:
        comparison_sheet = workbook["Comparison"]
        assert comparison_sheet.max_row == 2
        assert comparison_sheet["A2"].value == "shotcut"
        assert comparison_sheet["B2"].value == "demo"
        assert comparison_sheet["C2"].value == "demo_baseline"
        assert comparison_sheet["D2"].value == "demo_turbo"
        assert comparison_sheet["E2"].value == 10.0
        assert comparison_sheet["F2"].value == 5.0
        assert comparison_sheet["G2"].value == '=IF(OR(E2="",F2="",E2<=0),"",(E2-F2)/E2)'
    finally:
        workbook.close()
