from __future__ import annotations

import csv
import json
from argparse import Namespace
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

import full_test_pipeline
import xlsx_report_generator


def _write_monitor_csv(csv_path: Path, *, software: str, test_name: str, duration_seconds: float) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=("software", "duration_seconds", "started_at", "ended_at", "test_name"),
        )
        writer.writeheader()
        writer.writerow(
            {
                "software": software,
                "duration_seconds": str(duration_seconds),
                "started_at": "2026-04-21T10:00:00.000+08:00",
                "ended_at": "2026-04-21T10:00:10.000+08:00",
                "test_name": test_name,
            }
        )


class DummyLauncher:
    def __init__(self) -> None:
        self.launched: list[str] = []
        self.waited: list[tuple[str, float]] = []

    def launch(self, software: str) -> None:
        self.launched.append(software)

    def wait_for_main_window(self, software: str, timeout: float = 30.0) -> None:
        self.waited.append((software, timeout))


class DummyOperator:
    def __init__(self, software: str, event_log: list[tuple[str, str]]) -> None:
        self.software = software
        self.event_log = event_log

    def perform(self, input_video_path: Path, output_video_path: Path) -> None:
        assert input_video_path.exists()
        output_video_path.parent.mkdir(parents=True, exist_ok=True)
        output_video_path.write_bytes(b"fake-video")
        self.event_log.append(("perform", self.software))

    def close(self) -> None:
        self.event_log.append(("close", self.software))


class DummyMonitorBridge:
    def __init__(self) -> None:
        self.sessions: dict[str, tuple[str, str, Path]] = {}
        self.waited_sessions: list[str] = []
        self.stopped_sessions: list[str] = []

    def start_background_monitor(self, software: str, test_name: str, output_path: Path):
        output_path.parent.mkdir(parents=True, exist_ok=True)
        session_id = f"{software}-{test_name}"
        self.sessions[session_id] = (software, test_name, output_path)
        session_dir = output_path.parent.parent / "runtime" / session_id
        session_dir.mkdir(parents=True, exist_ok=True)
        return SimpleNamespace(
            session_id=session_id,
            output_path=output_path,
            state_path=session_dir / "state.json",
            session_output_path=session_dir / "session.csv",
            worker_stdout_path=session_dir / "worker.stdout.log",
            worker_stderr_path=session_dir / "worker.stderr.log",
            status="running",
        )

    def wait_for_session_completion(self, session_id: str) -> dict[str, str]:
        self.waited_sessions.append(session_id)
        software, test_name, output_path = self.sessions[session_id]
        duration_seconds = 5.0 if test_name.endswith("turbo") else 10.0
        _write_monitor_csv(
            output_path,
            software=software,
            test_name=test_name,
            duration_seconds=duration_seconds,
        )
        return {"status": "completed"}

    def stop_session(self, session_id: str, *, wait_seconds: float = 15.0) -> dict[str, str]:
        self.stopped_sessions.append(session_id)
        software, test_name, output_path = self.sessions[session_id]
        duration_seconds = 4.0 if test_name.endswith("turbo") else 9.0
        _write_monitor_csv(
            output_path,
            software=software,
            test_name=test_name,
            duration_seconds=duration_seconds,
        )
        return {"status": "completed_with_warnings"}

    def run_blender_monitor(
        self,
        *,
        test_name: str,
        output_path: Path,
        blend_file: Path,
        blender_executable: Path | None = None,
        blender_ui_mode: str = "visible",
        render_mode: str = "frame",
        frame: int = 1,
    ) -> Path:
        assert blend_file.exists()
        assert blender_ui_mode in {"visible", "headless"}
        duration_seconds = 5.0 if test_name.endswith("turbo") else 10.0
        _write_monitor_csv(
            output_path,
            software="blender",
            test_name=test_name,
            duration_seconds=duration_seconds,
        )
        session_dir = output_path.parent.parent / "runtime" / f"blender-{test_name}"
        session_dir.mkdir(parents=True, exist_ok=True)
        return SimpleNamespace(
            session_id=f"blender-{test_name}",
            output_path=output_path,
            state_path=session_dir / "state.json",
            session_output_path=session_dir / "session.csv",
            worker_stdout_path=None,
            worker_stderr_path=None,
            status="completed",
        )


class DummyAiTurboController:
    def __init__(self) -> None:
        self.configured: list[str] = []
        self.started = 0
        self.stopped = 0
        self.sequence_calls: list[tuple[str, float]] = []
        self.ensure_running_calls = 0
        self.close_calls = 0

    def start_topmost_guard(self) -> None:
        self.started += 1

    def stop_topmost_guard(self) -> None:
        self.stopped += 1

    def ensure_running(self) -> None:
        self.ensure_running_calls += 1

    def configure_for_software(self, software: str) -> None:
        self.configured.append(software)

    def run_sequence(self, software: str, *, wait_seconds: float) -> None:
        self.sequence_calls.append((software, wait_seconds))

    def close(self) -> None:
        self.close_calls += 1


def test_build_pipeline_cases_expands_shotcut_for_4k_big_workload(tmp_path: Path) -> None:
    big_input = tmp_path / "4K_big.mp4"
    small_input = tmp_path / "4K_small.mp4"
    big_input.write_bytes(b"big")
    small_input.write_bytes(b"small")

    cases = full_test_pipeline.build_pipeline_cases(
        ("shotcut", "kdenlive"),
        [("baseline", False)],
        workload_name="demo",
        input_video_path=big_input,
    )

    assert [case.case_id for case in cases] == [
        "shotcut__baseline",
        "shotcut__4k_1big_1small__baseline",
        "shotcut__4k_2big_2small__baseline",
        "kdenlive__baseline",
        "kdenlive__4k_1big_1small__baseline",
        "kdenlive__4k_2big_2small__baseline",
    ]
    assert [case.sequence_order for case in cases[:3]] == [1, 2, 3]
    assert [case.case_slug for case in cases[:3]] == [
        "",
        full_test_pipeline.SHOTCUT_MULTI_CASE_SMALL_SLUG,
        full_test_pipeline.SHOTCUT_MULTI_CASE_DOUBLE_SLUG,
    ]
    assert [case.sequence_name for case in cases[:3]] == [full_test_pipeline.SHOTCUT_SEQUENCE_NAME] * 3
    assert [case.sequence_order for case in cases[3:6]] == [1, 2, 3]
    assert [case.sequence_name for case in cases[3:6]] == [full_test_pipeline.KDENLIVE_SEQUENCE_NAME] * 3


def test_build_pipeline_cases_keeps_default_layout_for_non_4k_big_input(tmp_path: Path) -> None:
    input_video = tmp_path / "input.mp4"
    input_video.write_bytes(b"video")

    cases = full_test_pipeline.build_pipeline_cases(
        ("shotcut", "kdenlive"),
        [("baseline", False)],
        workload_name="demo",
        input_video_path=input_video,
    )

    assert [case.case_id for case in cases] == [
        "shotcut__baseline",
        "kdenlive__baseline",
    ]


def test_run_shotcut_case_sequence_reuses_one_session_for_three_cases(monkeypatch, tmp_path: Path) -> None:
    big_input = tmp_path / "4K_big.mp4"
    small_input = tmp_path / "4K_small.mp4"
    big_input.write_bytes(b"big")
    small_input.write_bytes(b"small")

    pipeline_paths = full_test_pipeline.build_pipeline_paths(tmp_path / "results", "demo")
    recorder = full_test_pipeline.PipelineRunRecorder(pipeline_paths.root_dir)
    launcher = DummyLauncher()
    monitor_bridge = DummyMonitorBridge()
    window = object()
    steps: list[tuple[str, object]] = []

    class SequenceShotcutOperator:
        def close(self) -> None:
            steps.append(("close", None))

        def _connect_main_window(self, timeout=30.0):
            steps.append(("connect", timeout))
            return window

        def _open_input_clip(self, current_window, input_path: Path):
            assert current_window is window or current_window is None
            steps.append(("open_input", input_path.name))
            return window

        def _append_selected_clip_to_timeline(self, current_window) -> None:
            assert current_window is window
            steps.append(("append", None))

        def _export_current_timeline(self, current_window, output_path: Path):
            assert current_window is window
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_bytes(b"video")
            steps.append(("export", output_path.name))
            return window

    monkeypatch.setattr(full_test_pipeline, "build_operator", lambda software: SequenceShotcutOperator())

    cases = full_test_pipeline.build_pipeline_cases(
        ("shotcut",),
        [("baseline", False)],
        workload_name="demo",
        input_video_path=big_input,
    )

    results = full_test_pipeline.run_shotcut_case_sequence(
        cases=cases,
        workload_name="demo",
        input_video_path=big_input,
        pipeline_paths=pipeline_paths,
        launcher=launcher,
        monitor_bridge=monitor_bridge,
        ai_turbo_controller=None,
        recorder=recorder,
    )

    assert launcher.launched == ["shotcut"]
    assert launcher.waited == [("shotcut", 30.0)]
    assert [name for name, _ in steps] == [
        "close",
        "connect",
        "open_input",
        "append",
        "export",
        "open_input",
        "append",
        "export",
        "append",
        "open_input",
        "append",
        "export",
        "close",
    ]
    assert steps[2:5] == [("open_input", "4K_big.mp4"), ("append", None), ("export", "shotcut_demo_baseline.mp4")]
    assert steps[5:8] == [
        ("open_input", "4K_small.mp4"),
        ("append", None),
        ("export", "shotcut_4k_1big_1small_baseline.mp4"),
    ]
    assert steps[8:12] == [
        ("append", None),
        ("open_input", "4K_big.mp4"),
        ("append", None),
        ("export", "shotcut_4k_2big_2small_baseline.mp4"),
    ]
    assert [result.case_id for result, _ in results] == [
        "shotcut__baseline",
        "shotcut__4k_1big_1small__baseline",
        "shotcut__4k_2big_2small__baseline",
    ]
    assert [monitor_path.name for _, monitor_path in [(case, result.csv_path) for case, result in results if result is not None]] == [
        "shotcut_demo_baseline.csv",
        "shotcut_4k_1big_1small_baseline.csv",
        "shotcut_4k_2big_2small_baseline.csv",
    ]


def test_run_kdenlive_case_sequence_reuses_one_session_for_three_cases(monkeypatch, tmp_path: Path) -> None:
    big_input = tmp_path / "4K_big.mp4"
    small_input = tmp_path / "4K_small.mp4"
    big_input.write_bytes(b"big")
    small_input.write_bytes(b"small")

    pipeline_paths = full_test_pipeline.build_pipeline_paths(tmp_path / "results", "demo")
    recorder = full_test_pipeline.PipelineRunRecorder(pipeline_paths.root_dir)
    launcher = DummyLauncher()
    monitor_bridge = DummyMonitorBridge()
    window = object()
    steps: list[tuple[str, object]] = []

    class SequenceKdenliveOperator:
        def close(self) -> None:
            steps.append(("close", None))

        def _connect_main_window(self, timeout=30.0):
            steps.append(("connect", timeout))
            return window

        def _open_input_clip(self, current_window, input_path: Path):
            assert current_window is window or current_window is None
            steps.append(("open_input", input_path.name))
            return window

        def _insert_clip_to_timeline(self, current_window) -> None:
            assert current_window is window
            steps.append(("insert", None))

        def _close_render_dialog_only(self) -> None:
            steps.append(("close_render_dialog", None))

        def _render_current_timeline(self, current_window, output_path: Path):
            assert current_window is window
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_bytes(b"video")
            steps.append(("render", output_path.name))
            return window

    monkeypatch.setattr(full_test_pipeline, "build_operator", lambda software: SequenceKdenliveOperator())

    cases = tuple(
        case
        for case in full_test_pipeline.build_pipeline_cases(
            ("kdenlive",),
            [("baseline", False)],
            workload_name="demo",
            input_video_path=big_input,
        )
        if case.software == "kdenlive"
    )

    results = full_test_pipeline.run_kdenlive_case_sequence(
        cases=cases,
        workload_name="demo",
        input_video_path=big_input,
        pipeline_paths=pipeline_paths,
        launcher=launcher,
        monitor_bridge=monitor_bridge,
        ai_turbo_controller=None,
        recorder=recorder,
    )

    assert launcher.launched == ["kdenlive"]
    assert launcher.waited == [("kdenlive", 30.0)]
    assert [name for name, _ in steps] == [
        "close",
        "connect",
        "open_input",
        "insert",
        "render",
        "open_input",
        "insert",
        "render",
        "close_render_dialog",
        "insert",
        "open_input",
        "insert",
        "render",
        "close",
    ]
    assert [case.case_id for case, _ in results] == [
        "kdenlive__baseline",
        "kdenlive__4k_1big_1small__baseline",
        "kdenlive__4k_2big_2small__baseline",
    ]
    assert [result.csv_path.name for case, result in results if result is not None] == [
        "kdenlive_demo_baseline.csv",
        "kdenlive_4k_1big_1small_baseline.csv",
        "kdenlive_4k_2big_2small_baseline.csv",
    ]


def test_run_pipeline_with_test_doubles_generates_comparison_report(monkeypatch, tmp_path: Path) -> None:
    input_video = tmp_path / "input.mp4"
    input_video.write_bytes(b"input-video")
    blend_file = tmp_path / "scene.blend"
    blend_file.write_text("blend", encoding="utf-8")

    launcher = DummyLauncher()
    monitor_bridge = DummyMonitorBridge()
    ai_turbo_controller = DummyAiTurboController()
    operator_log: list[tuple[str, str]] = []

    monkeypatch.setattr(full_test_pipeline, "SoftwareLauncher", lambda: launcher)
    monkeypatch.setattr(full_test_pipeline, "MonitorBridge", lambda: monitor_bridge)
    monkeypatch.setattr(full_test_pipeline, "AiTurboEngineController", lambda **kwargs: ai_turbo_controller)
    monkeypatch.setattr(full_test_pipeline, "resolve_input_video", lambda workload_name: input_video)
    monkeypatch.setattr(
        full_test_pipeline,
        "build_operator",
        lambda software: DummyOperator(software, operator_log),
    )
    monkeypatch.setattr(
        xlsx_report_generator,
        "collect_device_info",
        lambda: [("device_name", "pytest-host"), ("windows_system_version", "pytest-os")],
    )

    args = Namespace(
        workload_name="demo",
        input_video="",
        software=["shotcut", "blender"],
        results_root=str(tmp_path / "results"),
        ai_turbo_window_title=full_test_pipeline.DEFAULT_AI_TURBO_TITLE_RE,
        ai_turbo_shortcut="",
        blend_file=str(blend_file),
        blender_exe="",
        blender_ui_mode="visible",
        render_mode="frame",
        frame=1,
        resume_run_root="",
        skip_baseline=False,
        skip_turbo=False,
    )

    report_path = full_test_pipeline.run_pipeline(args)
    assert report_path.exists()
    run_root = report_path.parent.parent
    progress_log = run_root / "pipeline_progress.jsonl"
    summary_path = run_root / "pipeline_summary.md"
    manifest_path = run_root / "pipeline_manifest.json"
    assert progress_log.exists()
    assert summary_path.exists()
    assert manifest_path.exists()
    summary_text = summary_path.read_text(encoding="utf-8")
    assert "status: completed" in summary_text
    assert "shotcut__baseline" in summary_text
    assert "blender__turbo" in summary_text
    assert "case_state.json" in summary_text

    shotcut_case_state = run_root / "cases" / "shotcut__baseline" / "case_state.json"
    blender_case_state = run_root / "cases" / "blender__turbo" / "case_state.json"
    shotcut_state = json.loads(shotcut_case_state.read_text(encoding="utf-8"))
    blender_state = json.loads(blender_case_state.read_text(encoding="utf-8"))
    assert shotcut_state["status"] == "completed"
    assert shotcut_state["session_id"] == "shotcut-demo_baseline"
    assert shotcut_state["traceback_path"] == ""
    assert shotcut_state["log_path"].endswith("case_events.jsonl")
    assert blender_state["status"] == "completed"
    assert blender_state["session_id"] == "blender-demo_turbo"

    workbook = load_workbook(report_path, data_only=False)
    try:
        comparison_sheet = workbook["Comparison"]
        assert comparison_sheet.max_row == 3
        assert {comparison_sheet["A2"].value, comparison_sheet["A3"].value} == {"shotcut", "blender"}
        assert comparison_sheet["E2"].value == 10.0
        assert comparison_sheet["F2"].value == 5.0
        assert comparison_sheet["E3"].value == 10.0
        assert comparison_sheet["F3"].value == 5.0
        assert comparison_sheet["G2"].value == '=IF(OR(E2="",F2="",E2<=0),"",(E2-F2)/E2)'
        assert comparison_sheet["G3"].value == '=IF(OR(E3="",F3="",E3<=0),"",(E3-F3)/E3)'
    finally:
        workbook.close()

    assert ai_turbo_controller.started == 1
    assert ai_turbo_controller.stopped == 1
    assert ai_turbo_controller.configured == ["shotcut", "blender"]
    assert ("perform", "shotcut") in operator_log


def test_build_pipeline_paths_resolves_relative_results_root(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.chdir(tmp_path)

    pipeline_paths = full_test_pipeline.build_pipeline_paths(Path("results/pipeline_runs"), "demo")

    expected_root = (tmp_path / "results" / "pipeline_runs").resolve(strict=False)
    assert pipeline_paths.root_dir.is_absolute()
    assert pipeline_paths.csv_dir.is_absolute()
    assert pipeline_paths.export_dir.is_absolute()
    assert pipeline_paths.report_dir.is_absolute()
    assert pipeline_paths.root_dir.parent == expected_root
    assert pipeline_paths.csv_dir.parent == pipeline_paths.root_dir
    assert pipeline_paths.export_dir.parent == pipeline_paths.root_dir
    assert pipeline_paths.report_dir.parent == pipeline_paths.root_dir


def test_run_pipeline_honors_excluded_software(monkeypatch, tmp_path: Path) -> None:
    input_video = tmp_path / "input.mp4"
    input_video.write_bytes(b"input-video")
    selected_software_per_pass: list[tuple[str, ...]] = []

    def _fake_run_pipeline_pass(**kwargs):
        selected_software_per_pass.append(kwargs["selected_software"])
        return []

    def _fake_generate_xlsx_report(csv_paths, report_path, default_test_case):
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_bytes(b"fake-report")
        return len(csv_paths), len(csv_paths), report_path

    monkeypatch.setattr(full_test_pipeline, "resolve_input_video", lambda workload_name: input_video)
    monkeypatch.setattr(full_test_pipeline, "run_pipeline_pass", _fake_run_pipeline_pass)
    monkeypatch.setattr(full_test_pipeline, "SoftwareLauncher", lambda: object())
    monkeypatch.setattr(full_test_pipeline, "MonitorBridge", lambda: object())
    monkeypatch.setattr(full_test_pipeline, "AiTurboEngineController", lambda **kwargs: object())
    monkeypatch.setattr(full_test_pipeline, "generate_xlsx_report", _fake_generate_xlsx_report)
    monkeypatch.setattr(
        full_test_pipeline,
        "collect_completed_case_csv_paths",
        lambda pipeline_paths, cases: [pipeline_paths.csv_dir / "kdenlive_demo_baseline.csv"],
    )

    args = Namespace(
        workload_name="demo",
        input_video="",
        software=["shotcut", "kdenlive", "blender"],
        exclude_software=["shotcut"],
        results_root=str(tmp_path / "results"),
        ai_turbo_window_title=full_test_pipeline.DEFAULT_AI_TURBO_TITLE_RE,
        ai_turbo_shortcut="",
        blend_file="",
        blender_exe="",
        blender_ui_mode="visible",
        render_mode="frame",
        frame=1,
        resume_run_root="",
        skip_baseline=False,
        skip_turbo=True,
    )

    report_path = full_test_pipeline.run_pipeline(args)

    assert report_path.exists()
    assert selected_software_per_pass == [("kdenlive", "blender")]


def test_run_ai_turbo_sequence_uses_explicit_software_and_wait(monkeypatch) -> None:
    ai_turbo_controller = DummyAiTurboController()

    monkeypatch.setattr(full_test_pipeline, "AiTurboEngineController", lambda **kwargs: ai_turbo_controller)

    args = Namespace(
        software=["shotcut", "kdenlive"],
        exclude_software=[],
        ai_turbo_window_title=full_test_pipeline.DEFAULT_AI_TURBO_TITLE_RE,
        ai_turbo_shortcut="",
        ai_turbo_sequence_software="kdenlive",
        ai_turbo_sequence_wait_seconds=12.5,
    )

    software = full_test_pipeline.run_ai_turbo_sequence(args)

    assert software == "kdenlive"
    assert ai_turbo_controller.sequence_calls == [("kdenlive", 12.5)]


def test_run_ai_turbo_open_check_ensures_running_and_closes(monkeypatch) -> None:
    ai_turbo_controller = DummyAiTurboController()

    monkeypatch.setattr(full_test_pipeline, "AiTurboEngineController", lambda **kwargs: ai_turbo_controller)
    monkeypatch.setattr(full_test_pipeline.time, "sleep", lambda _seconds: None)

    args = Namespace(
        ai_turbo_window_title=full_test_pipeline.DEFAULT_AI_TURBO_TITLE_RE,
        ai_turbo_shortcut="",
        ai_turbo_open_check_wait_seconds=12.5,
    )

    full_test_pipeline.run_ai_turbo_open_check(args)

    assert ai_turbo_controller.ensure_running_calls == 1
    assert ai_turbo_controller.close_calls == 1
    assert ai_turbo_controller.sequence_calls == []


def test_run_ai_turbo_sequence_infers_single_selected_software(monkeypatch) -> None:
    ai_turbo_controller = DummyAiTurboController()

    monkeypatch.setattr(full_test_pipeline, "AiTurboEngineController", lambda **kwargs: ai_turbo_controller)

    args = Namespace(
        software=["shotcut"],
        exclude_software=[],
        ai_turbo_window_title=full_test_pipeline.DEFAULT_AI_TURBO_TITLE_RE,
        ai_turbo_shortcut="",
        ai_turbo_sequence_software="",
        ai_turbo_sequence_wait_seconds=0.0,
    )

    software = full_test_pipeline.run_ai_turbo_sequence(args)

    assert software == "shotcut"
    assert ai_turbo_controller.sequence_calls == [("shotcut", 0.0)]


def test_resolve_ai_turbo_sequence_software_requires_single_selection() -> None:
    args = Namespace(
        software=["shotcut", "kdenlive"],
        exclude_software=[],
        ai_turbo_sequence_software="",
    )

    with pytest.raises(AssertionError, match="Standalone AI Turbo sequence needs exactly one target software"):
        full_test_pipeline.resolve_ai_turbo_sequence_software(args)


@pytest.mark.parametrize("software", ["avidemux", "kdenlive"])
def test_run_non_blender_case_stops_monitor_after_automation(monkeypatch, tmp_path: Path, software: str) -> None:
    input_video = tmp_path / "input.mp4"
    input_video.write_bytes(b"input-video")
    pipeline_paths = full_test_pipeline.build_pipeline_paths(tmp_path / "results", "demo")
    launcher = DummyLauncher()
    monitor_bridge = DummyMonitorBridge()
    operator_log: list[tuple[str, str]] = []
    recorder = full_test_pipeline.PipelineRunRecorder(pipeline_paths.root_dir)

    monkeypatch.setattr(
        full_test_pipeline,
        "build_operator",
        lambda current_software: DummyOperator(current_software, operator_log),
    )

    result = full_test_pipeline.run_non_blender_case(
        software=software,
        variant="baseline",
        workload_name="demo",
        input_video_path=input_video,
        requested_csv_path=full_test_pipeline.build_requested_csv_path(
            pipeline_paths,
            software=software,
            workload_name="demo",
            variant="baseline",
        ),
        output_video_path=full_test_pipeline.build_export_output_path(
            pipeline_paths,
            software=software,
            workload_name="demo",
            variant="baseline",
        ),
        pipeline_paths=pipeline_paths,
        launcher=launcher,
        monitor_bridge=monitor_bridge,
        ai_turbo_controller=None,
        recorder=recorder,
    )

    assert result.csv_path.exists()
    assert monitor_bridge.stopped_sessions == [f"{software}-demo_baseline"]
    assert monitor_bridge.waited_sessions == []


def test_build_export_output_path_uses_mkv_for_avidemux(tmp_path: Path) -> None:
    pipeline_paths = full_test_pipeline.build_pipeline_paths(tmp_path / "results", "demo")

    output_path = full_test_pipeline.build_export_output_path(
        pipeline_paths,
        software="avidemux",
        workload_name="demo",
        variant="baseline",
    )

    assert output_path.suffix == ".mkv"


def test_run_pipeline_resume_reruns_only_failed_cases(monkeypatch, tmp_path: Path) -> None:
    input_video = tmp_path / "input.mp4"
    input_video.write_bytes(b"input-video")

    launcher = DummyLauncher()
    monitor_bridge = DummyMonitorBridge()
    perform_counts: dict[str, int] = {"shotcut": 0, "kdenlive": 0}

    class FailOnceOperator(DummyOperator):
        def perform(self, input_video_path: Path, output_video_path: Path) -> None:
            perform_counts[self.software] += 1
            if self.software == "kdenlive" and perform_counts[self.software] == 1:
                raise RuntimeError("kdenlive failed once")
            super().perform(input_video_path, output_video_path)

    monkeypatch.setattr(full_test_pipeline, "SoftwareLauncher", lambda: launcher)
    monkeypatch.setattr(full_test_pipeline, "MonitorBridge", lambda: monitor_bridge)
    monkeypatch.setattr(full_test_pipeline, "AiTurboEngineController", lambda **kwargs: DummyAiTurboController())
    monkeypatch.setattr(full_test_pipeline, "resolve_input_video", lambda workload_name: input_video)
    monkeypatch.setattr(full_test_pipeline, "build_operator", lambda software: FailOnceOperator(software, []))
    monkeypatch.setattr(
        xlsx_report_generator,
        "collect_device_info",
        lambda: [("device_name", "pytest-host"), ("windows_system_version", "pytest-os")],
    )

    first_args = Namespace(
        workload_name="demo",
        input_video="",
        software=["shotcut", "kdenlive"],
        exclude_software=[],
        results_root=str(tmp_path / "results"),
        ai_turbo_window_title=full_test_pipeline.DEFAULT_AI_TURBO_TITLE_RE,
        ai_turbo_shortcut="",
        blend_file="",
        blender_exe="",
        blender_ui_mode="visible",
        render_mode="frame",
        frame=1,
        resume_run_root="",
        skip_baseline=False,
        skip_turbo=True,
    )

    with pytest.raises(RuntimeError, match="kdenlive failed once"):
        full_test_pipeline.run_pipeline(first_args)

    run_root = next((tmp_path / "results").resolve(strict=False).glob("demo_*"))
    shotcut_state_path = run_root / "cases" / "shotcut__baseline" / "case_state.json"
    kdenlive_state_path = run_root / "cases" / "kdenlive__baseline" / "case_state.json"
    shotcut_state = json.loads(shotcut_state_path.read_text(encoding="utf-8"))
    kdenlive_state = json.loads(kdenlive_state_path.read_text(encoding="utf-8"))
    assert shotcut_state["status"] == "completed"
    assert kdenlive_state["status"] == "failed"

    resume_args = Namespace(
        workload_name="demo",
        input_video="",
        software=["shotcut", "kdenlive"],
        exclude_software=[],
        results_root=str(tmp_path / "results"),
        ai_turbo_window_title=full_test_pipeline.DEFAULT_AI_TURBO_TITLE_RE,
        ai_turbo_shortcut="",
        blend_file="",
        blender_exe="",
        blender_ui_mode="visible",
        render_mode="frame",
        frame=1,
        resume_run_root=str(run_root),
        skip_baseline=False,
        skip_turbo=True,
    )

    report_path = full_test_pipeline.run_pipeline(resume_args)

    assert report_path.exists()
    assert perform_counts == {"shotcut": 1, "kdenlive": 2}
    resumed_shotcut_state = json.loads(shotcut_state_path.read_text(encoding="utf-8"))
    resumed_kdenlive_state = json.loads(kdenlive_state_path.read_text(encoding="utf-8"))
    assert resumed_shotcut_state["attempt_count"] == 1
    assert resumed_kdenlive_state["attempt_count"] == 2
    assert resumed_kdenlive_state["status"] == "completed"
    assert resumed_kdenlive_state["traceback_path"] == ""
